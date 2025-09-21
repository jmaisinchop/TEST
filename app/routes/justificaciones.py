from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from sqlalchemy.orm import joinedload
from sqlalchemy import or_
from collections import defaultdict
from app import db
from app.models import Justificaciones, PersonnelEmployee, PersonnelDepartment
from datetime import datetime
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill

justificaciones_bp = Blueprint('justificaciones', __name__, url_prefix='/justificaciones')


def get_filtered_justificaciones(q, depto_ids, fecha_desde_str, fecha_hasta_str):
    """Función auxiliar para obtener justificaciones filtradas, usada por index y descarga."""
    query = Justificaciones.query.join(
        PersonnelEmployee, Justificaciones.employee_passport == PersonnelEmployee.passport
    ).join(
        PersonnelDepartment, PersonnelEmployee.department_id == PersonnelDepartment.id
    ).filter(Justificaciones.anulada == False)

    if q:
        search_term = f'%{q}%'
        query = query.filter(or_(
            PersonnelEmployee.first_name.ilike(search_term),
            PersonnelEmployee.last_name.ilike(search_term),
            PersonnelEmployee.passport.ilike(search_term)
        ))

    # Nuevo filtro para múltiples departamentos
    if depto_ids:
        query = query.filter(PersonnelDepartment.id.in_(depto_ids))

    if fecha_desde_str:
        query = query.filter(Justificaciones.date_end >= fecha_desde_str)
    if fecha_hasta_str:
        query = query.filter(Justificaciones.date_start <= fecha_hasta_str)

    return query.order_by(Justificaciones.date_start.desc()).all()


@justificaciones_bp.route('/')
def index():
    q = request.args.get('q', '', type=str)
    # Se obtienen los IDs como una lista de enteros
    depto_ids = request.args.getlist('depto_id', type=int)
    fecha_desde_str = request.args.get('desde', '', type=str)
    fecha_hasta_str = request.args.get('hasta', '', type=str)

    justificaciones_filtradas = get_filtered_justificaciones(q, depto_ids, fecha_desde_str, fecha_hasta_str)

    agrupado = defaultdict(list)
    for j in justificaciones_filtradas:
        clave_empleado = (
            j.employee_passport,
            f"{j.empleado.first_name} {j.empleado.last_name}",
            j.empleado.department.dept_name
        )
        agrupado[clave_empleado].append(j)

    departamentos_permitidos = ['Callcenter', 'Guayaquil', 'Administracion']
    departamentos = PersonnelDepartment.query.filter(
        PersonnelDepartment.dept_name.in_(departamentos_permitidos)
    ).order_by(PersonnelDepartment.dept_name).all()

    return render_template(
        'justificaciones/index.html',
        agrupado=agrupado,
        departamentos=departamentos,
        q=q,
        depto_ids_seleccionados=depto_ids,  # Se pasa la lista de IDs seleccionados
        desde=fecha_desde_str,
        hasta=fecha_hasta_str
    )


@justificaciones_bp.route('/descargar-reporte')
def descargar_reporte():
    """Genera y descarga un reporte en Excel de las justificaciones filtradas."""
    q = request.args.get('q', '', type=str)
    depto_ids = request.args.getlist('depto_id', type=int)
    fecha_desde_str = request.args.get('desde', '', type=str)
    fecha_hasta_str = request.args.get('hasta', '', type=str)

    justificaciones = get_filtered_justificaciones(q, depto_ids, fecha_desde_str, fecha_hasta_str)

    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reporte de Justificaciones"

    headers = ["Cédula", "Empleado", "Departamento", "Tipo", "Fecha Inicio", "Fecha Fin", "Razón"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")

    for j in justificaciones:
        sheet.append([
            j.employee_passport,
            f"{j.empleado.first_name} {j.empleado.last_name}",
            j.empleado.department.dept_name,
            j.justification_type.replace('_', ' ').title(),
            j.date_start,
            j.date_end,
            j.reason
        ])

    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 50 else 50
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save(output)
    output.seek(0)

    return send_file(
        output, as_attachment=True, download_name='reporte_justificaciones.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# --- El resto de las rutas (crear, editar, anular, etc.) no cambian ---
@justificaciones_bp.route('/crear', methods=['POST'])
def crear():
    passport = request.form.get('employee_passport')
    jtype = request.form.get('justification_type')
    dstart_str = request.form.get('date_start')
    dend_str = request.form.get('date_end')
    reason = request.form.get('reason', '')

    if not all([passport, jtype, dstart_str, dend_str]):
        flash('Todos los campos son obligatorios, excepto la razón.', 'warning')
        return redirect(url_for('justificaciones.index'))

    if not PersonnelEmployee.query.filter_by(passport=passport).first():
        flash(f'El empleado con pasaporte "{passport}" no fue encontrado.', 'danger')
        return redirect(url_for('justificaciones.index'))

    dstart_obj = datetime.strptime(dstart_str, '%Y-%m-%d').date()
    dend_obj = datetime.strptime(dend_str, '%Y-%m-%d').date()

    conflicto = Justificaciones.query.filter(
        Justificaciones.employee_passport == passport,
        Justificaciones.anulada == False,
        Justificaciones.date_start <= dend_obj,
        Justificaciones.date_end >= dstart_obj
    ).first()

    if conflicto:
        flash(
            f'Error: Ya existe una justificación para este empleado que se cruza con las fechas seleccionadas (del {conflicto.date_start.strftime("%d-%m-%Y")} al {conflicto.date_end.strftime("%d-%m-%Y")}).',
            'danger')
        return redirect(url_for('justificaciones.index'))

    nueva_justificacion = Justificaciones(
        employee_passport=passport, justification_type=jtype,
        date_start=dstart_obj, date_end=dend_obj, reason=reason
    )
    db.session.add(nueva_justificacion)
    db.session.commit()
    flash('Justificación creada exitosamente.', 'success')
    return redirect(url_for('justificaciones.index'))


@justificaciones_bp.route('/cargar-excel', methods=['POST'])
def cargar_excel():
    if 'archivo_excel' not in request.files:
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('justificaciones.index'))

    file = request.files['archivo_excel']
    if file.filename == '':
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('justificaciones.index'))

    if file and file.filename.endswith('.xlsx'):
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            creadas = 0;
            errores = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                passport, jtype, dstart, dend, reason = row
                if not all([passport, jtype, dstart, dend]):
                    errores += 1;
                    continue
                if not PersonnelEmployee.query.filter_by(passport=str(passport)).first():
                    errores += 1;
                    continue

                dstart_obj = dstart.date() if isinstance(dstart, datetime) else datetime.strptime(
                    str(dstart).split(" ")[0], '%Y-%m-%d').date()
                dend_obj = dend.date() if isinstance(dend, datetime) else datetime.strptime(str(dend).split(" ")[0],
                                                                                            '%Y-%m-%d').date()

                conflicto = Justificaciones.query.filter(Justificaciones.employee_passport == str(passport),
                                                         Justificaciones.anulada == False,
                                                         Justificaciones.date_start <= dend_obj,
                                                         Justificaciones.date_end >= dstart_obj).first()
                if conflicto:
                    errores += 1;
                    continue

                nueva_justificacion = Justificaciones(employee_passport=str(passport), justification_type=str(jtype),
                                                      date_start=dstart_obj, date_end=dend_obj,
                                                      reason=str(reason or ''))
                db.session.add(nueva_justificacion)
                creadas += 1

            db.session.commit()
            flash(f'Carga masiva completada: {creadas} justificaciones creadas, {errores} filas con errores.',
                  'success' if errores == 0 else 'warning')
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error al procesar el archivo: {e}', 'danger')
        return redirect(url_for('justificaciones.index'))
    flash('Formato de archivo no válido. Por favor, sube un archivo .xlsx', 'danger')
    return redirect(url_for('justificaciones.index'))


@justificaciones_bp.route('/descargar-plantilla')
def descargar_plantilla():
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Plantilla de Justificaciones"
    headers = ["Pasaporte", "Tipo_Justificacion", "Fecha_Inicio", "Fecha_Fin", "Razon"]
    sheet.append(headers)
    sheet.append(["123456789", "vacaciones", "2025-12-20", "2025-12-22", "Vacaciones de fin de año"])
    sheet.append(["987654321", "incapacidad_medica", "2025-11-10", "2025-11-12", "Reposo médico"])
    sheet.cell(row=5, column=1, value="Tipos de justificación válidos:").font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=6, column=1, value="permiso_personal");
    sheet.cell(row=7, column=1, value="incapacidad_medica");
    sheet.cell(row=8, column=1, value="vacaciones");
    sheet.cell(row=9, column=1, value="calamidad_domestica")
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='plantilla_justificaciones.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@justificaciones_bp.route('/editar/<int:justificacion_id>', methods=['POST'])
def editar(justificacion_id):
    justificacion = db.get_or_404(Justificaciones, justificacion_id)
    justificacion.justification_type = request.form.get('justification_type')
    justificacion.date_start = datetime.strptime(request.form.get('date_start'), '%Y-%m-%d').date()
    justificacion.date_end = datetime.strptime(request.form.get('date_end'), '%Y-%m-%d').date()
    justificacion.reason = request.form.get('reason', '')
    db.session.commit()
    flash('Justificación actualizada correctamente.', 'success')
    return redirect(url_for('justificaciones.index'))


@justificaciones_bp.route('/anular/<int:justificacion_id>', methods=['POST'])
def anular(justificacion_id):
    justificacion = db.get_or_404(Justificaciones, justificacion_id)
    justificacion.anulada = True
    db.session.commit()
    flash('Justificación anulada correctamente.', 'success')
    return redirect(url_for('justificaciones.index'))


@justificaciones_bp.route('/buscar_empleados')
def buscar_empleados():
    search_term = request.args.get('term', '')
    departamentos_permitidos = ['Callcenter', 'Administracion', 'Guayaquil']
    query_obj = PersonnelEmployee.query.join(PersonnelDepartment).filter(
        PersonnelDepartment.dept_name.in_(departamentos_permitidos),
        or_(
            PersonnelEmployee.first_name.ilike(f'%{search_term}%'),
            PersonnelEmployee.last_name.ilike(f'%{search_term}%'),
            PersonnelEmployee.passport.ilike(f'%{search_term}%')
        )
    ).limit(10)
    empleados = query_obj.all()
    resultados = [
        {'label': f'{emp.first_name} {emp.last_name} ({emp.passport})', 'value': emp.passport}
        for emp in empleados
    ]
    return jsonify(resultados)
