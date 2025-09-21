from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from app import db
from app.models import (Carteras, Grupos, GrupoEmpleados, GrupoHorariosEspeciales, PersonnelEmployee)
from datetime import datetime, time
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill
from collections import defaultdict

# Creamos el nuevo Blueprint
asignacion_masiva_bp = Blueprint('asignacion_masiva', __name__, url_prefix='/asignacion-masiva')


@asignacion_masiva_bp.route('/')
def index():
    """Muestra la página de la herramienta de carga masiva."""
    return render_template('asignacion_masiva/index.html')


@asignacion_masiva_bp.route('/descargar-plantilla')
def descargar_plantilla():
    """Genera y envía un archivo Excel que sirve como plantilla de carga."""
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Plantilla de Carga Masiva"

    headers = [
        "Pasaporte_Empleado", "Nombre_Cartera", "Nombre_Grupo",
        "Fecha_Horario_Especial (Opcional)", "Entrada_Especial (Opcional)",
        "Salida_Especial (Opcional)", "Horas_Extras (Opcional)"
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")

    sheet.append([
        "123456789", "Campaña A", "Equipo Alpha", "2025-12-24", "10:00", "16:00", 0
    ])
    sheet.append([
        "987654321", "Campaña B", "Equipo Beta", "2025-11-15", "", "", 4
    ])
    sheet.append([
        "112233445", "Campaña A", "Equipo Alpha", "", "", "", ""
    ])

    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='plantilla_carga_masiva.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def parse_time_from_excel(value):
    """Función robusta para convertir un valor de Excel a un objeto time."""
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, str):
        for fmt in ('%H:%M:%S', '%H:%M'):
            try:
                return datetime.strptime(value, fmt).time()
            except ValueError:
                continue
    return None


@asignacion_masiva_bp.route('/procesar-excel', methods=['POST'])
def procesar_excel():
    """Procesa el archivo Excel subido para realizar las asignaciones masivas."""
    if 'archivo_excel' not in request.files:
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('asignacion_masiva.index'))

    file = request.files['archivo_excel']
    if file.filename == '':
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('asignacion_masiva.index'))

    if not file.filename.endswith('.xlsx'):
        flash('Formato de archivo no válido. Por favor, sube un archivo .xlsx', 'danger')
        return redirect(url_for('asignacion_masiva.index'))

    carteras_cache = {c.name: c for c in Carteras.query.all()}
    grupos_cache = {g.name: g for g in Grupos.query.all()}
    empleados_cache = {e.passport: e for e in PersonnelEmployee.query.all()}

    stats = defaultdict(int)
    error_details = []

    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            passport, cartera_nombre, grupo_nombre, fecha_he, entrada_he, salida_he, horas_extras = row

            if not all([passport, cartera_nombre, grupo_nombre]):
                stats['filas_con_errores'] += 1
                error_details.append(f"Fila {row_idx}: Datos básicos incompletos.")
                continue

            passport = str(passport)
            if passport not in empleados_cache:
                stats['empleados_no_encontrados'] += 1
                stats['filas_con_errores'] += 1
                error_details.append(f"Fila {row_idx}: Pasaporte '{passport}' no encontrado.")
                continue

            # --- ✨ LÓGICA MEJORADA PARA GESTIÓN DE CARTERAS ---
            cartera = carteras_cache.get(cartera_nombre)
            # Si la cartera no existe O si existe pero está anulada, se crea una nueva.
            if not cartera or cartera.anulada:
                last_cartera = Carteras.query.order_by(Carteras.id.desc()).first()
                new_id = (last_cartera.id + 1) if last_cartera else 1
                code = f"CAR-{new_id:04d}"

                # Se crea una nueva cartera con un nuevo código.
                cartera = Carteras(name=cartera_nombre, code=code)
                db.session.add(cartera)
                db.session.flush()  # Para obtener el ID de la nueva cartera

                # Se actualiza el caché para esta sesión, reemplazando la anulada si existía.
                carteras_cache[cartera_nombre] = cartera
                stats['carteras_creadas'] += 1

            # --- Gestión de Grupos (Crear si no existe) ---
            grupo = grupos_cache.get(grupo_nombre)
            if not grupo:
                last_grupo = Grupos.query.order_by(Grupos.id.desc()).first()
                new_id = (last_grupo.id + 1) if last_grupo else 1
                code = f"GRP-{new_id:04d}"
                grupo = Grupos(name=grupo_nombre, code=code, cartera_id=cartera.id, hora_entrada=time(8, 0),
                               hora_salida=time(18, 0))
                db.session.add(grupo)
                db.session.flush()
                grupos_cache[grupo_nombre] = grupo
                stats['grupos_creados'] += 1

            # --- Asignación de Empleado a Grupo ---
            asignacion_existente = GrupoEmpleados.query.filter_by(grupo_id=grupo.id, employee_passport=passport).first()
            if not asignacion_existente:
                nueva_asignacion = GrupoEmpleados(grupo_id=grupo.id, employee_passport=passport)
                db.session.add(nueva_asignacion)
                stats['asignaciones_creadas'] += 1

            # --- Gestión de Horarios Especiales ---
            if fecha_he:
                try:
                    fecha_obj = fecha_he.date() if isinstance(fecha_he, datetime) else datetime.strptime(
                        str(fecha_he).split(" ")[0], '%Y-%m-%d').date()

                    horario_especial = GrupoHorariosEspeciales.query.filter_by(grupo_id=grupo.id,
                                                                               fecha=fecha_obj).first()
                    if not horario_especial:
                        horario_especial = GrupoHorariosEspeciales(grupo_id=grupo.id, fecha=fecha_obj)
                        db.session.add(horario_especial)

                    if entrada_he:
                        horario_especial.hora_entrada_especial = parse_time_from_excel(entrada_he)
                    if salida_he:
                        horario_especial.hora_salida_especial = parse_time_from_excel(salida_he)
                    if horas_extras is not None and str(horas_extras).isdigit():
                        horario_especial.horas_extras = int(horas_extras)

                    stats['horarios_especiales_creados_o_actualizados'] += 1
                except (ValueError, TypeError) as e:
                    stats['filas_con_errores'] += 1
                    error_details.append(f"Fila {row_idx}: Formato de fecha/hora inválido - {e}")
                    continue

        db.session.commit()

        success_message = (
            f"Proceso completado: {stats['carteras_creadas']} carteras nuevas, {stats['grupos_creados']} grupos nuevos, "
            f"{stats['asignaciones_creadas']} empleados asignados, {stats['horarios_especiales_creados_o_actualizados']} horarios especiales aplicados."
        )
        flash(success_message, 'success')

        if error_details:
            error_message = f"{stats['filas_con_errores']} filas tuvieron errores. Detalles: " + " | ".join(
                error_details[:3])  # Muestra los primeros 3 errores
            flash(error_message, 'danger')

    except Exception as e:
        db.session.rollback()
        flash(f'Ocurrió un error crítico al procesar el archivo: {e}', 'danger')

    return redirect(url_for('asignacion_masiva.index'))
