from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from sqlalchemy import or_
from collections import defaultdict
from app import db
from app.models import Permisos, PersonnelEmployee, PersonnelDepartment
from datetime import datetime
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill

permisos_bp = Blueprint('permisos', __name__, url_prefix='/permisos')


def get_filtered_permisos(q, depto_ids, fecha_desde_str, fecha_hasta_str):
    """Función auxiliar para obtener permisos filtrados."""
    query = Permisos.query.join(
        PersonnelEmployee, Permisos.employee_passport == PersonnelEmployee.passport
    ).join(
        PersonnelDepartment, PersonnelEmployee.department_id == PersonnelDepartment.id
    )

    if q:
        search_term = f'%{q}%'
        query = query.filter(or_(
            PersonnelEmployee.first_name.ilike(search_term),
            PersonnelEmployee.last_name.ilike(search_term),
            PersonnelEmployee.passport.ilike(search_term)
        ))

    if depto_ids:
        query = query.filter(PersonnelDepartment.id.in_(depto_ids))

    if fecha_desde_str:
        query = query.filter(Permisos.fecha >= fecha_desde_str)
    if fecha_hasta_str:
        query = query.filter(Permisos.fecha <= fecha_hasta_str)

    return query.order_by(Permisos.fecha.desc()).all()


@permisos_bp.route('/')
def index():
    q = request.args.get('q', '', type=str)
    depto_ids = request.args.getlist('depto_id', type=int)
    fecha_desde_str = request.args.get('desde', '', type=str)
    fecha_hasta_str = request.args.get('hasta', '', type=str)

    permisos_filtrados = get_filtered_permisos(q, depto_ids, fecha_desde_str, fecha_hasta_str)

    agrupado = defaultdict(list)
    for p in permisos_filtrados:
        clave_empleado = (
            p.employee_passport,
            f"{p.empleado.first_name} {p.empleado.last_name}",
            p.empleado.department.dept_name
        )
        agrupado[clave_empleado].append(p)

    departamentos_permitidos = ['Callcenter', 'Guayaquil', 'Administracion']
    departamentos = PersonnelDepartment.query.filter(
        PersonnelDepartment.dept_name.in_(departamentos_permitidos)
    ).order_by(PersonnelDepartment.dept_name).all()

    return render_template(
        'permisos/index.html',
        agrupado=agrupado,
        departamentos=departamentos,
        q=q,
        depto_ids_seleccionados=depto_ids,
        desde=fecha_desde_str,
        hasta=fecha_hasta_str
    )


@permisos_bp.route('/descargar-reporte')
def descargar_reporte():
    q = request.args.get('q', '', type=str)
    depto_ids = request.args.getlist('depto_id', type=int)
    fecha_desde_str = request.args.get('desde', '', type=str)
    fecha_hasta_str = request.args.get('hasta', '', type=str)

    permisos = get_filtered_permisos(q, depto_ids, fecha_desde_str, fecha_hasta_str)

    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reporte de Permisos"

    headers = ["Cédula", "Empleado", "Departamento", "Fecha", "Desde Hora", "Hasta Hora", "Motivo", "Observación"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True);
        cell.fill = PatternFill("solid", fgColor="DDEBF7")

    for p in permisos:
        sheet.append([
            p.employee_passport, f"{p.empleado.first_name} {p.empleado.last_name}",
            p.empleado.department.dept_name, p.fecha, p.hora_desde, p.hora_hasta,
            p.motivo, p.observacion
        ])

    for col in sheet.columns:
        max_length = 0;
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 50 else 50
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='reporte_permisos.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@permisos_bp.route('/crear', methods=['POST'])
def crear():
    passport = request.form.get('employee_passport')
    fecha_str = request.form.get('fecha')
    hora_desde_str = request.form.get('hora_desde')
    hora_hasta_str = request.form.get('hora_hasta')
    motivo = request.form.get('motivo')
    observacion = request.form.get('observacion', '')

    if not all([passport, fecha_str, hora_desde_str, hora_hasta_str, motivo]):
        flash('Todos los campos son obligatorios.', 'warning')
        return redirect(url_for('permisos.index'))

    if not PersonnelEmployee.query.filter_by(passport=passport).first():
        flash(f'El empleado con pasaporte "{passport}" no fue encontrado.', 'danger')
        return redirect(url_for('permisos.index'))

    nuevo_permiso = Permisos(
        employee_passport=passport,
        fecha=datetime.strptime(fecha_str, '%Y-%m-%d').date(),
        hora_desde=datetime.strptime(hora_desde_str, '%H:%M').time(),
        hora_hasta=datetime.strptime(hora_hasta_str, '%H:%M').time(),
        motivo=motivo, observacion=observacion
    )
    db.session.add(nuevo_permiso)
    db.session.commit()
    flash('Permiso creado exitosamente.', 'success')
    return redirect(url_for('permisos.index'))


@permisos_bp.route('/cargar-excel', methods=['POST'])
def cargar_excel():
    if 'archivo_excel' not in request.files:
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('permisos.index'))
    file = request.files['archivo_excel']
    if file.filename == '':
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('permisos.index'))

    if file and file.filename.endswith('.xlsx'):
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            creadas = 0;
            errores = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                passport, fecha, hora_desde, hora_hasta, motivo, observacion = row
                if not all([passport, fecha, hora_desde, hora_hasta, motivo]):
                    errores += 1;
                    continue
                if not PersonnelEmployee.query.filter_by(passport=str(passport)).first():
                    errores += 1;
                    continue

                fecha_obj = fecha.date() if isinstance(fecha, datetime) else datetime.strptime(str(fecha).split(" ")[0],
                                                                                               '%Y-%m-%d').date()
                hora_desde_obj = hora_desde if isinstance(hora_desde, datetime.time) else datetime.strptime(
                    str(hora_desde), '%H:%M:%S').time()
                hora_hasta_obj = hora_hasta if isinstance(hora_hasta, datetime.time) else datetime.strptime(
                    str(hora_hasta), '%H:%M:%S').time()

                nuevo_permiso = Permisos(
                    employee_passport=str(passport), fecha=fecha_obj, hora_desde=hora_desde_obj,
                    hora_hasta=hora_hasta_obj, motivo=str(motivo), observacion=str(observacion or '')
                )
                db.session.add(nuevo_permiso)
                creadas += 1

            db.session.commit()
            flash(f'Carga masiva completada: {creadas} permisos creados, {errores} filas con errores.',
                  'success' if errores == 0 else 'warning')
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error al procesar el archivo: {e}', 'danger')
        return redirect(url_for('permisos.index'))
    flash('Formato de archivo no válido. Por favor, sube un archivo .xlsx', 'danger')
    return redirect(url_for('permisos.index'))


@permisos_bp.route('/descargar-plantilla')
def descargar_plantilla():
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Plantilla de Permisos"
    headers = ["Pasaporte", "Fecha", "Hora_Desde", "Hora_Hasta", "Motivo", "Observacion"]
    sheet.append(headers)
    sheet.append(["123456789", "2025-12-24", "14:00:00", "18:00:00", "Trámite personal", ""])
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='plantilla_permisos.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@permisos_bp.route('/editar/<int:permiso_id>', methods=['POST'])
def editar(permiso_id):
    permiso = db.get_or_404(Permisos, permiso_id)
    permiso.fecha = datetime.strptime(request.form.get('fecha'), '%Y-%m-%d').date()
    permiso.hora_desde = datetime.strptime(request.form.get('hora_desde'), '%H:%M').time()
    permiso.hora_hasta = datetime.strptime(request.form.get('hora_hasta'), '%H:%M').time()
    permiso.motivo = request.form.get('motivo')
    permiso.observacion = request.form.get('observacion', '')
    db.session.commit()
    flash('Permiso actualizado correctamente.', 'success')
    return redirect(url_for('permisos.index'))


@permisos_bp.route('/eliminar/<int:permiso_id>', methods=['POST'])
def eliminar(permiso_id):
    permiso = db.get_or_404(Permisos, permiso_id)
    db.session.delete(permiso)
    db.session.commit()
    flash('Permiso eliminado correctamente.', 'success')
    return redirect(url_for('permisos.index'))


@permisos_bp.route('/buscar_empleados')
def buscar_empleados():
    search_term = request.args.get('term', '')
    departamentos_permitidos = ['Callcenter', 'Administracion', 'Guayaquil']
    query_obj = PersonnelEmployee.query.join(PersonnelDepartment).filter(
        PersonnelDepartment.dept_name.in_(departamentos_permitidos),
        or_(
            PersonnelEmployee.first_name.ilike(f'%{search_term}%'),
            PersonnelEmployee.last_name.ilike(f'%{search_term}%'),
            PersonnelEmployee.passport.ilike(f'%{search_term}%')
        )
    ).limit(10)
    empleados = query_obj.all()
    resultados = [{'label': f'{emp.first_name} {emp.last_name} ({emp.passport})', 'value': emp.passport} for emp in
                  empleados]
    return jsonify(resultados)
