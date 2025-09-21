# --- IMPORTS NECESARIOS PARA AMBAS FUNCIONES ---
from collections import defaultdict
from datetime import timedelta, datetime, time
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text

# --- IMPORTACIONES DE LA APLICACIÓN (EJEMPLO) ---
# Asegúrate de que estas importaciones coincidan con la estructura de tu proyecto.
from app import db
from app.models import (PersonnelEmployee, PersonnelDepartment, Justificaciones,
                        IClockTransaction, Permisos, GrupoHorariosEspeciales, DepartmentHorariosEspeciales,
                        Grupos, GrupoEmpleados)


# ==============================================================================
# --- FUNCIÓN 1: MOTOR DE CÁLCULO DEL REPORTE (SIN CAMBIOS) ---
# ==============================================================================

def build_report(start_date, end_date, department_id=None):
    """
    Construye el reporte de asistencia con todos los cálculos detallados.
    """
    departamentos_permitidos = ['Callcenter', 'Guayaquil', 'Administracion']
    query_empleados = PersonnelEmployee.query.join(PersonnelDepartment).filter(
        PersonnelDepartment.dept_name.in_(departamentos_permitidos)
    )
    if department_id and department_id.isdigit():
        query_empleados = query_empleados.filter(PersonnelDepartment.id == int(department_id))
    empleados_a_reportar = query_empleados.order_by(PersonnelEmployee.last_name).all()
    if not empleados_a_reportar: return []

    pasaportes = [e.passport for e in empleados_a_reportar]
    ids_empleados = [e.id for e in empleados_a_reportar]

    sql = text("""
        SELECT p.passport, (t.punch_time AT TIME ZONE 'America/Guayaquil')::date AS fecha_local,
               (t.punch_time AT TIME ZONE 'America/Guayaquil')::time AS hora_local
        FROM iclock_transaction t JOIN personnel_employee p ON t.emp_id = p.id
        WHERE t.emp_id = ANY(:ids_empleados) AND t.punch_time >= :start_date AND t.punch_time < :end_date_plus_one
    """)
    marcaciones_q = db.session.execute(sql, {"ids_empleados": ids_empleados, "start_date": start_date,
                                             "end_date_plus_one": end_date + timedelta(days=1)}).fetchall()

    justificaciones_q = Justificaciones.query.filter(Justificaciones.employee_passport.in_(pasaportes),
                                                     Justificaciones.anulada == False,
                                                     Justificaciones.date_start <= end_date,
                                                     Justificaciones.date_end >= start_date).all()
    permisos_q = Permisos.query.filter(Permisos.employee_passport.in_(pasaportes),
                                       Permisos.fecha.between(start_date, end_date)).all()
    grupo_horarios_q = GrupoHorariosEspeciales.query.filter(
        GrupoHorariosEspeciales.fecha.between(start_date, end_date)).all()
    depto_horarios_q = DepartmentHorariosEspeciales.query.filter(
        DepartmentHorariosEspeciales.fecha.between(start_date, end_date)).all()

    empleado_grupo_map = {ge.employee_passport: ge.grupo_id for ge, g in
                          db.session.query(GrupoEmpleados, Grupos).join(Grupos).all()}
    marcaciones_map = defaultdict(lambda: defaultdict(list))
    for m in marcaciones_q: marcaciones_map[m.passport][m.fecha_local.strftime('%Y-%m-%d')].append(m.hora_local)
    justificaciones_map = defaultdict(list)
    for j in justificaciones_q:
        d = j.date_start
        while d <= j.date_end: justificaciones_map[j.employee_passport].append(d); d += timedelta(days=1)
    permisos_map = defaultdict(dict)
    for p in permisos_q: permisos_map[p.employee_passport][p.fecha] = p

    horarios_especiales_map = defaultdict(dict)
    for dh in depto_horarios_q:
        horarios_especiales_map[('depto', dh.dept_name, dh.fecha)] = {'extras': dh.horas_extras, 'feriado': dh.feriado,
                                                                      'entrada': dh.hora_entrada_especial,
                                                                      'salida': dh.hora_salida_especial}
    for gh in grupo_horarios_q:
        horarios_especiales_map[('grupo', gh.grupo_id, gh.fecha)] = {'extras': gh.horas_extras, 'feriado': gh.feriado,
                                                                     'entrada': gh.hora_entrada_especial,
                                                                     'salida': gh.hora_salida_especial}

    reporte_final = []
    dias_del_periodo = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1) if
                        (start_date + timedelta(days=i)).weekday() != 6]

    for empleado in empleados_a_reportar:
        resumen = {
            'total_asistencias': 0,
            'total_atrasos_normal': 0, 'total_atrasos_sabfer': 0,
            'total_minutos_atraso_normal': 0, 'total_minutos_atraso_sabfer': 0,
            'total_faltas_normal': 0, 'total_faltas_sabfer': 0,
            'total_faltas_injustificadas_normal': 0, 'total_faltas_injustificadas_sabfer': 0,
            'total_faltas_justificadas': 0,
            'total_horas_extras_normal': timedelta(), 'total_horas_extras_sabfer': timedelta()
        }

        registros_diarios_emp = []
        for dia in dias_del_periodo:
            grupo_id = empleado_grupo_map.get(empleado.passport)
            regla_depto = horarios_especiales_map.get(('depto', empleado.department.dept_name, dia), {})
            regla_grupo = horarios_especiales_map.get(('grupo', grupo_id, dia), {})
            horas_extras_aprobadas_int = max(regla_depto.get('extras', 0), regla_grupo.get('extras', 0))
            es_feriado = regla_depto.get('feriado', False) or regla_grupo.get('feriado', False)
            horario_entrada_prog = regla_grupo.get('entrada') or regla_depto.get('entrada') or time(8, 0, 0)
            horario_salida_prog = regla_grupo.get('salida') or regla_depto.get('salida') or time(18, 0, 0)
            permiso_del_dia = permisos_map[empleado.passport].get(dia)
            if permiso_del_dia:
                if permiso_del_dia.hora_desde <= horario_entrada_prog: horario_entrada_prog = permiso_del_dia.hora_hasta
                if permiso_del_dia.hora_hasta >= horario_salida_prog: horario_salida_prog = permiso_del_dia.hora_desde
            hora_limite_entrada = (datetime.combine(dia, horario_entrada_prog) + timedelta(minutes=5)).time()

            registro = {
                'fecha': dia, 'estado': '-',
                'horario_prog': f"{horario_entrada_prog.strftime('%H:%M')} - {horario_salida_prog.strftime('%H:%M')}",
                'marcaciones': '-', 'minutos_atraso': 0, 'tiempo_trabajado': '00:00',
                'horas_extras_trabajadas': '00:00', 'es_falta': 0, 'es_atraso': 0,
                'hora_ingreso': None, 'hora_salida_almuerzo': None,
                'hora_regreso_almuerzo': None, 'hora_salida_final': None,
                'tiempo_almuerzo': '00:00'
            }

            if es_feriado and not horas_extras_aprobadas_int > 0:
                registro['estado'] = 'Feriado'
            elif dia in justificaciones_map[empleado.passport]:
                registro['estado'] = 'Justificado'
                resumen['total_faltas_justificadas'] += 1
            elif sorted(marcaciones_map[empleado.passport][dia.strftime('%Y-%m-%d')]):
                resumen['total_asistencias'] += 1
                marcaciones_dia = sorted(marcaciones_map[empleado.passport][dia.strftime('%Y-%m-%d')])
                if len(marcaciones_dia) >= 1: registro['hora_ingreso'] = marcaciones_dia[0]
                if len(marcaciones_dia) >= 2: registro['hora_salida_final'] = marcaciones_dia[-1]
                if len(marcaciones_dia) == 4:
                    registro['hora_salida_almuerzo'] = marcaciones_dia[1]
                    registro['hora_regreso_almuerzo'] = marcaciones_dia[2]
                    almuerzo_delta = datetime.combine(dia, marcaciones_dia[2]) - datetime.combine(dia, marcaciones_dia[1])
                    h_a, rem_a = divmod(int(almuerzo_delta.total_seconds()), 3600)
                    m_a, _ = divmod(rem_a, 60)
                    registro['tiempo_almuerzo'] = f'{h_a:02d}:{m_a:02d}'

                entrada_real = registro['hora_ingreso']
                if entrada_real > hora_limite_entrada:
                    atraso_delta = datetime.combine(dia, entrada_real) - datetime.combine(dia, hora_limite_entrada)
                    registro['minutos_atraso'] = int(atraso_delta.total_seconds() // 60)
                    registro['estado'] = 'Atraso'
                    registro['es_atraso'] = 1
                    resumen['total_atrasos_normal'] += 1
                    resumen['total_minutos_atraso_normal'] += registro['minutos_atraso']
                else:
                    registro['estado'] = 'Presente'

                if len(marcaciones_dia) >= 2:
                    salida_real = registro['hora_salida_final']
                    lunch_duration = timedelta()
                    if registro['hora_salida_almuerzo'] and registro['hora_regreso_almuerzo']:
                        lunch_duration = datetime.combine(dia, registro['hora_regreso_almuerzo']) - datetime.combine(dia, registro['hora_salida_almuerzo'])
                    duracion_neta = (datetime.combine(dia, salida_real) - datetime.combine(dia, entrada_real)) - lunch_duration
                    if duracion_neta.total_seconds() < 0: duracion_neta = timedelta(seconds=0)
                    h, rem = divmod(int(duracion_neta.total_seconds()), 3600)
                    m, _ = divmod(rem, 60)
                    registro['tiempo_trabajado'] = f'{h:02d}:{m:02d}'
            else:
                es_laborable = (dia.weekday() < 5 and not es_feriado) or (es_feriado and horas_extras_aprobadas_int > 0)
                if es_laborable:
                    registro['estado'] = 'Falta'
                    registro['es_falta'] = 1
                    resumen['total_faltas_normal'] += 1
                    resumen['total_faltas_injustificadas_normal'] += 1
                else:
                    registro['estado'] = 'Fin de Semana'
            registros_diarios_emp.append(registro)
        reporte_final.append({'empleado': empleado, 'registros': registros_diarios_emp, 'resumen': resumen})

    return reporte_final

# ==============================================================================
# --- FUNCIÓN 2: GENERADOR DE REPORTE EXCEL (MODIFICADO) ---
# ==============================================================================
def crear_excel_reporte(report_data, params):
    """
    Construye un libro de Excel elegante y profesional.
    Esta versión solo genera la hoja "Matriz de Trabajo".
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    # --- 1. CONFIGURACIÓN DE ESTILOS Y COLORES ---
    COLORS = {
        'primary_dark': '44546A', 'primary_light': 'DDEBF7', 'accent_green': 'C6E0B4',
        'accent_yellow': 'FFF2CC', 'accent_red': 'F8CBAD', 'accent_blue': 'B4C6E7',
        'font_light': 'FFFFFF', 'font_dark': '000000', 'border_grey': 'BFBFBF',
        'alt_row_fill': 'F2F2F2'
    }
    STYLES = {
        'title': Font(name='Calibri', size=16, bold=True, color=COLORS['primary_dark']),
        'subtitle': Font(name='Calibri', size=11, bold=True, color=COLORS['primary_dark']),
        'header': Font(name='Calibri', size=11, bold=True, color=COLORS['font_light']),
        'header_fill_resumen': PatternFill(start_color=COLORS['primary_dark'], fill_type="solid"),
        'header_fill_trabajo': PatternFill(start_color=COLORS['primary_light'], fill_type="solid"),
        'header_font_dark': Font(name='Calibri', size=11, bold=True, color=COLORS['font_dark']),
        'alt_row_fill': PatternFill(start_color=COLORS['alt_row_fill'], fill_type="solid"),
        'center_align': Alignment(horizontal='center', vertical='center'),
        'highlight_falta': PatternFill(start_color=COLORS['accent_red'], fill_type="solid"),
        'highlight_atraso': PatternFill(start_color=COLORS['accent_yellow'], fill_type="solid"),
        'highlight_justificado': PatternFill(start_color=COLORS['accent_blue'], fill_type="solid"),
    }
    THIN_BORDER = Border(left=Side(style='thin', color=COLORS['border_grey']),
                         right=Side(style='thin', color=COLORS['border_grey']),
                         top=Side(style='thin', color=COLORS['border_grey']),
                         bottom=Side(style='thin', color=COLORS['border_grey']))

    # --- 2. FUNCIONES AUXILIARES DE DISEÑO ---
    def set_report_header(sheet, title, params):
        sheet.cell(row=1, column=1, value=title).font = STYLES['title']
        sheet.cell(row=2, column=1,
                   value=f"Período del {params['fecha_desde'].strftime('%d-%m-%Y')} al {params['fecha_hasta'].strftime('%d-%m-%Y')}").font = \
        STYLES['subtitle']
        sheet.row_dimensions[1].height = 20

    def apply_table_style(sheet, start_row, end_row, start_col, end_col, header_fill, header_font=STYLES['header']):
        for cell in sheet[start_row][start_col - 1:end_col]:
            cell.fill = header_fill; cell.font = header_font
            cell.border = THIN_BORDER; cell.alignment = STYLES['center_align']
        for row_idx in range(start_row + 1, end_row + 1):
            is_alt_row = (row_idx - start_row) % 2 != 0
            for col_idx in range(start_col, end_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.border = THIN_BORDER
                if is_alt_row: cell.fill = STYLES['alt_row_fill']

    # --- 3. PREPARACIÓN DE DATOS COMUNES ---
    dias_semana_map = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

    # --- 4. CREACIÓN DE LA ÚNICA HOJA REQUERIDA ---

    # --- Hoja: Matriz de Trabajo (ÚNICA HOJA ACTIVA) ---
    sheet_trabajo = workbook.create_sheet(title="Matriz de Trabajo")
    set_report_header(sheet_trabajo, "Matriz Detallada de Trabajo", params)
    sheet_trabajo.append([])

    headers_trabajo = [
        "Cédula", "Empleado", "Departamento", "Fecha", "Día", "Estado",
        "Tiene Atraso", "Entrada Programada", "Entrada Real", "Minutos Atraso",
        "H. Salida Alm.", "H. Reg. Alm.", "H. Salida Final", "T. Almuerzo", "T. Trabajado"
    ]
    sheet_trabajo.append(headers_trabajo)
    start_row_trabajo = 5

    for item in report_data:
        empleado = item['empleado']
        nombre_completo = f"{empleado.last_name} {empleado.first_name}"
        for reg in item['registros']:
            fecha_dia = reg['fecha']
            estado = reg.get('estado', '')
            tiene_atraso = 'Sí' if reg.get('es_atraso') == 1 else 'No'
            entrada_programada = reg.get('horario_prog', ' - ').split(' - ')[0] if reg.get('estado') not in ['Falta', 'Feriado', 'Justificado', 'Fin de Semana'] else ''
            minutos_atraso = reg.get('minutos_atraso', 0) if reg.get('es_atraso') == 1 else ''
            row_data = [
                empleado.passport, nombre_completo, empleado.department.dept_name,
                fecha_dia.strftime('%Y-%m-%d'), dias_semana_map[fecha_dia.weekday()], estado
            ]
            row_data.extend([
                tiene_atraso, entrada_programada, reg.get('hora_ingreso'), minutos_atraso
            ])
            if estado not in ['Falta', 'Justificado', 'Feriado', 'Fin de Semana']:
                row_data.extend([
                    reg.get('hora_salida_almuerzo'), reg.get('hora_regreso_almuerzo'),
                    reg.get('hora_salida_final'), reg.get('tiempo_almuerzo', '00:00'),
                    reg.get('tiempo_trabajado', '00:00')
                ])
            else:
                row_data.extend(['-'] * 5)
            sheet_trabajo.append(row_data)

    if sheet_trabajo.max_row >= start_row_trabajo:
        apply_table_style(sheet_trabajo, start_row_trabajo - 1, sheet_trabajo.max_row, 1, len(headers_trabajo),
                          STYLES['header_fill_trabajo'], STYLES['header_font_dark'])
        for row in sheet_trabajo.iter_rows(min_row=start_row_trabajo, min_col=6, max_col=6):
            for cell in row:
                if cell.value == 'Falta': cell.fill = STYLES['highlight_falta']
                elif cell.value == 'Atraso': cell.fill = STYLES['highlight_atraso']
                elif cell.value == 'Justificado': cell.fill = STYLES['highlight_justificado']
    sheet_trabajo.freeze_panes = 'A5'


    # --- OTRAS HOJAS (DESACTIVADAS MEDIANTE COMENTARIOS) ---

    # --- Hoja 1 DESACTIVADA: Resumen General ---
    # sheet_resumen = workbook.create_sheet(title="Resumen General")
    # ... (código para la hoja de resumen) ...

    # --- Hoja 3 DESACTIVADA: Matriz de Asistencia ---
    # asistencia_codes = {'Presente': 'P', 'Atraso': 'A', 'Falta': 'FI', 'Justificado': 'FJ'}
    # asistencia_highlights = {'FI': STYLES['highlight_falta'], 'A': STYLES['highlight_atraso']}
    # crear_hoja_matriz("Matriz Asistencia", 'estado', asistencia_codes, asistencia_highlights)

    # --- Hojas 4, 5 y 6 DESACTIVADAS: Detalles de Faltas, Atrasos, Extras ---
    # ... (código para las hojas de detalle) ...


    # --- 5. AJUSTE FINAL DE COLUMNAS ---
    # Este bucle ahora solo afectará a la única hoja activa.
    for sheet in workbook.worksheets:
        for col_cells in sheet.columns:
            max_length = 0
            if isinstance(col_cells[0], openpyxl.cell.cell.MergedCell):
                continue
            for cell in col_cells:
                if cell.value:
                    try:
                        padding = 4 if cell.font and cell.font.bold else 0
                        cell_len = len(str(cell.value)) + padding
                        if cell_len > max_length: max_length = cell_len
                    except: pass
            adjusted_width = max(12, min(max_length + 2, 50))
            sheet.column_dimensions[get_column_letter(col_cells[0].column)].width = adjusted_width

    # --- 6. GUARDAR Y DEVOLVER EL ARCHIVO ---
    in_memory_file = BytesIO()
    workbook.save(in_memory_file)
    in_memory_file.seek(0)
    return in_memory_file